"""Excel export utilities for job data."""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import structlog
import os

from ..core.models import JobListing
from ..core.config import get_settings

logger = structlog.get_logger()


class ExcelExporter:
    """Handles exporting job data to Excel format."""
    
    def __init__(self):
        self.settings = get_settings()
        
    def export_jobs(self, jobs: List[JobListing], filename: Optional[str] = None, append_mode: bool = False) -> Path:
        """
        Export job listings to Excel file with Korean headers.
        
        Args:
            jobs: List of JobListing objects to export
            filename: Optional custom filename
            append_mode: Whether to append to existing file or create new one
            
        Returns:
            Path to the created Excel file
        """
        try:
            # Prepare data for DataFrame with Korean formatting
            job_data = [self._job_to_dict(job) for job in jobs]
            
            if not job_data:
                logger.warning("No job data to export")
                return None
            
            # Create DataFrame
            df = pd.DataFrame(job_data)
            
            # Define Korean column order - CORRECTED to match actual fields
            korean_column_order = [
                "채용제목", "회사명", "근무지역", "고용형태", "경력수준",
                "급여정보", "최소급여", "최대급여", "급여통화",
                "경력요구사항", "학력요구사항", "근무일정", "근무시간",
                "담당자명", "담당자연락처", "담당자이메일",
                "직무내용", "지원자격", "복리후생",
                "기업규모", "회사설명", "업종", "부서",
                "원격근무", "게시일", "마감일",
                "요구기술", "기술스택", "채용링크", "채용ID",
                "출처사이트", "수집일시"
            ]
            
            # Reorder columns (keep any extra columns at the end)
            available_columns = [col for col in korean_column_order if col in df.columns]
            extra_columns = [col for col in df.columns if col not in korean_column_order]
            df = df[available_columns + extra_columns]
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"채용공고_{timestamp}.xlsx"  # Korean filename
            
            # Ensure output directory exists
            self.settings.output_dir.mkdir(parents=True, exist_ok=True)
            
            # Create full file path
            file_path = self.settings.output_dir / filename
            
            # Handle append mode vs new file creation
            if append_mode and file_path.exists():
                # Append to existing file
                self._append_to_excel(file_path, df)
            else:
                # Create new file or overwrite existing
                with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                    df.to_excel(writer, sheet_name="채용공고", index=False)  # Korean sheet name
                    
                    # Get the workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets["채용공고"]
                    
                    # Apply formatting
                    self._format_excel_sheet(worksheet, df)
            
            logger.info(f"Exported {len(jobs)} jobs to {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"Failed to export jobs to Excel: {e}")
            raise
    
    def _job_to_dict(self, job: JobListing) -> Dict[str, Any]:
        """Convert JobListing to dictionary for DataFrame with Korean formatting."""
        job_dict = job.model_dump()
        
        # Create Korean-mapped dictionary
        korean_dict = {}
        
        # Map English fields to Korean with proper formatting - FIXED MAPPINGS
        field_mapping = {
            "title": ("채용제목", job_dict.get("title", "")),
            "company": ("회사명", job_dict.get("company", "")),
            "location": ("근무지역", job_dict.get("location", "")),
            "job_type": ("고용형태", self._format_job_type_korean(job_dict.get("job_type"))),
            "experience_level": ("경력수준", self._format_experience_level_korean(job_dict.get("experience_level"))),
            # Use actual salary fields from the model instead of non-existent ones
            "salary_min": ("최소급여", job_dict.get("salary_min", "")),
            "salary_max": ("최대급여", job_dict.get("salary_max", "")),
            "salary_currency": ("급여통화", job_dict.get("salary_currency", "")),
            # Map to actual fields that exist in JobListing model
            "requirements": ("지원자격", job_dict.get("requirements", "")),
            "benefits": ("복리후생", job_dict.get("benefits", "")),
            "description": ("직무내용", job_dict.get("description", "")),
            "recruiter_name": ("담당자명", job_dict.get("recruiter_name", "")),
            "recruiter_phone": ("담당자연락처", job_dict.get("recruiter_phone", "")),
            "recruiter_email": ("담당자이메일", job_dict.get("recruiter_email", "")),
            "remote_option": ("원격근무", self._format_boolean_korean(job_dict.get("remote_option"))),
            "posted_date": ("게시일", self._format_date_korean(job_dict.get("posted_date"))),
            "application_deadline": ("마감일", self._format_date_korean(job_dict.get("application_deadline"))),
            "skills": ("요구기술", self._format_list_korean(job_dict.get("skills"))),
            "technologies": ("기술스택", self._format_list_korean(job_dict.get("technologies"))),
            "company_size": ("기업규모", job_dict.get("company_size", "")),
            "company_description": ("회사설명", job_dict.get("company_description", "")),
            "department": ("부서", job_dict.get("department", "")),
            "industry": ("업종", job_dict.get("industry", "")),
            "job_url": ("채용링크", job_dict.get("job_url", "")),
            "job_id": ("채용ID", job_dict.get("job_id", "")),
            "source_site": ("출처사이트", job_dict.get("source_site", "")),
            "scraped_at": ("수집일시", self._format_datetime_korean(job_dict.get("scraped_at")))
        }
        
        # Build Korean dictionary
        for eng_key, (kor_key, value) in field_mapping.items():
            korean_dict[kor_key] = value if value else ""
        
        # Add Work24-specific additional data fields if they exist
        additional_data = job_dict.get("additional_data", {})
        if additional_data:
            # Map additional fields that might exist in additional_data
            if additional_data.get("salary_text"):
                korean_dict["급여정보"] = additional_data["salary_text"]
            if additional_data.get("education"):
                korean_dict["학력요구사항"] = additional_data["education"]
            if additional_data.get("work_schedule"):
                korean_dict["근무일정"] = additional_data["work_schedule"]
            if additional_data.get("work_hours"):
                korean_dict["근무시간"] = additional_data["work_hours"]
            if additional_data.get("experience"):
                korean_dict["경력요구사항"] = additional_data["experience"]
            if additional_data.get("job_type_korean"):
                korean_dict["고용형태상세"] = additional_data["job_type_korean"]
            if additional_data.get("work_schedule_details"):
                korean_dict["근무조건상세"] = additional_data["work_schedule_details"]
            if additional_data.get("application_method"):
                korean_dict["지원방법"] = additional_data["application_method"]
            if additional_data.get("required_documents"):
                korean_dict["제출서류"] = additional_data["required_documents"]
        
        return korean_dict
    
    def _format_job_type_korean(self, job_type) -> str:
        """Format job type in Korean."""
        if not job_type:
            return ""
        
        if hasattr(job_type, 'value'):
            job_type = job_type.value
        
        job_type_mapping = {
            "FULL_TIME": "정규직",
            "PART_TIME": "시간제",
            "CONTRACT": "계약직",
            "TEMPORARY": "임시직",
            "INTERNSHIP": "인턴십",
            "FREELANCE": "프리랜서",
            "정규직": "정규직",
            "시간제": "시간제"
        }
        
        return job_type_mapping.get(str(job_type), str(job_type))
    
    def _format_experience_level_korean(self, experience_level) -> str:
        """Format experience level in Korean."""
        if not experience_level:
            return ""
        
        if hasattr(experience_level, 'value'):
            experience_level = experience_level.value
        
        experience_mapping = {
            "ENTRY_LEVEL": "신입",
            "MID_LEVEL": "경력 3-5년",
            "SENIOR_LEVEL": "경력 5년 이상",
            "EXECUTIVE": "임원급",
            "INTERNSHIP": "인턴"
        }
        
        return experience_mapping.get(str(experience_level), str(experience_level))
    
    def _format_boolean_korean(self, value) -> str:
        """Format boolean values in Korean."""
        if value is None:
            return ""
        return "가능" if value else "불가능"
    
    def _format_date_korean(self, date_value) -> str:
        """Format date in Korean."""
        if not date_value:
            return ""
        if hasattr(date_value, 'strftime'):
            return date_value.strftime("%Y년 %m월 %d일")
        return str(date_value)
    
    def _format_datetime_korean(self, datetime_value) -> str:
        """Format datetime in Korean."""
        if not datetime_value:
            return ""
        if hasattr(datetime_value, 'strftime'):
            return datetime_value.strftime("%Y년 %m월 %d일 %H시 %M분")
        return str(datetime_value)
    
    def _format_list_korean(self, list_value) -> str:
        """Format list values in Korean."""
        if not list_value:
            return ""
        if isinstance(list_value, list):
            return ", ".join(str(item) for item in list_value)
        return str(list_value)
    
    def _format_excel_sheet(self, worksheet, df: pd.DataFrame) -> None:
        """Apply formatting to Excel sheet."""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Apply header formatting
        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for col_num, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_num)
            
            # Calculate optimal width
            max_length = max(
                len(str(column)),  # Header length
                max(len(str(df[column].iloc[i])) if i < len(df) else 0 
                    for i in range(min(len(df), 10)))  # Sample first 10 rows
            )
            
            # Set reasonable limits
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        worksheet.freeze_panes = "A2"
    
    def create_summary_report(self, jobs: List[JobListing]) -> Dict[str, Any]:
        """Create a summary report of the crawled jobs."""
        if not jobs:
            return {"total_jobs": 0}
        
        summary = {
            "total_jobs": len(jobs),
            "unique_companies": len(set(job.company for job in jobs)),
            "job_types": {},
            "experience_levels": {},
            "locations": {},
            "salary_stats": {},
            "top_skills": {},
            "sites_crawled": list(set(job.source_site for job in jobs))
        }
        
        # Job type distribution
        for job in jobs:
            if job.job_type:
                job_type_str = job.job_type.value.replace("_", " ").title()
                summary["job_types"][job_type_str] = summary["job_types"].get(job_type_str, 0) + 1
        
        # Experience level distribution
        for job in jobs:
            if job.experience_level:
                exp_str = job.experience_level.value.replace("_", " ").title()
                summary["experience_levels"][exp_str] = summary["experience_levels"].get(exp_str, 0) + 1
        
        # Location distribution (top 10)
        location_counts = {}
        for job in jobs:
            location_counts[job.location] = location_counts.get(job.location, 0) + 1
        summary["locations"] = dict(sorted(location_counts.items(), key=lambda x: x[1], reverse=True)[:10])
        
        # Salary statistics
        salaries = [job.salary_min for job in jobs if job.salary_min]
        if salaries:
            summary["salary_stats"] = {
                "min": float(min(salaries)),
                "max": float(max(salaries)),
                "avg": float(sum(salaries) / len(salaries))
            }
        
        # Top skills (top 15)
        skill_counts = {}
        for job in jobs:
            for skill in job.skills:
                skill_counts[skill] = skill_counts.get(skill, 0) + 1
        summary["top_skills"] = dict(sorted(skill_counts.items(), key=lambda x: x[1], reverse=True)[:15])
        
        return summary
    
    def _append_to_excel(self, file_path: Path, new_df: pd.DataFrame) -> None:
        """
        Append new data to existing Excel file, avoiding duplicates based on job_id.
        
        Args:
            file_path: Path to existing Excel file
            new_df: New data to append
        """
        try:
            # Read existing data
            existing_df = pd.read_excel(file_path, sheet_name="채용공고")
            
            # Filter out duplicates based on job_id (채용ID column)
            if "채용ID" in existing_df.columns and "채용ID" in new_df.columns:
                # Get existing job IDs
                existing_job_ids = set(existing_df["채용ID"].dropna().astype(str))
                
                # Filter new jobs to exclude duplicates
                before_count = len(new_df)
                new_df = new_df[~new_df["채용ID"].astype(str).isin(existing_job_ids)]
                after_count = len(new_df)
                
                duplicates_skipped = before_count - after_count
                if duplicates_skipped > 0:
                    print(f"🔍 {duplicates_skipped}개 중복 채용공고 제외됨")
                    logger.info(f"Skipped {duplicates_skipped} duplicate jobs based on job_id")
                
                # If no new unique jobs, don't append
                if len(new_df) == 0:
                    print("ℹ️ 새로운 채용공고가 없습니다 (모두 중복)")
                    logger.info("No new jobs to append - all were duplicates")
                    return
            
            # Combine with new data
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            
            # Write back to file with formatting
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                combined_df.to_excel(writer, sheet_name="채용공고", index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets["채용공고"]
                
                # Apply formatting
                self._format_excel_sheet(worksheet, combined_df)
            
            logger.info(f"Appended {len(new_df)} new jobs to existing file. Total: {len(combined_df)} jobs")
            
        except Exception as e:
            logger.error(f"Failed to append to Excel file: {e}")
            raise
    
    def save_incremental_batch(self, jobs: List[JobListing], batch_num: int, 
                              base_filename: Optional[str] = None, 
                              is_first_batch: bool = False) -> Path:
        """
        Save a batch of jobs incrementally to Excel.
        
        Args:
            jobs: List of JobListing objects to save
            batch_num: Batch number (for tracking)
            base_filename: Base filename to use (without .xlsx)
            is_first_batch: Whether this is the first batch (create new file)
            
        Returns:
            Path to the Excel file
        """
        try:
            # Generate base filename if not provided
            if not base_filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                base_filename = f"채용공고_{timestamp}"
            
            filename = f"{base_filename}.xlsx"
            
            if is_first_batch:
                # First batch - create new file
                logger.info(f"Creating new Excel file with batch {batch_num} ({len(jobs)} jobs)")
                print(f"📊 새 Excel 파일 생성 중... (배치 {batch_num}: {len(jobs)}개 채용공고)")
                return self.export_jobs(jobs, filename, append_mode=False)
            else:
                # Subsequent batches - append to existing file
                logger.info(f"Appending batch {batch_num} ({len(jobs)} jobs) to existing file")
                print(f"📊 Excel 파일에 추가 저장 중... (배치 {batch_num}: {len(jobs)}개 채용공고)")
                return self.export_jobs(jobs, filename, append_mode=True)
                
        except Exception as e:
            logger.error(f"Failed to save incremental batch {batch_num}: {e}")
            print(f"❌ 배치 {batch_num} 저장 실패: {e}")
            raise
    
